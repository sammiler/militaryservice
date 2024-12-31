from openpyxl import load_workbook
import sqlite3
# 读取Excel文件
wb = load_workbook('test.xlsx')

# 获取活动工作表（默认的第一个工作表）
sheet = wb.active

# 获取表格的列数
column_count = sheet.max_column
row_count = sheet.max_row
headers = []
for i in range(1, column_count + 1):
    for j in range(6, 3,-1):                #4-6行是表头
        cell = chr(65+ i - 1) + str(j)
        cell_value = sheet[cell].value
        if cell_value is not None:
            cell_value = cell_value.replace(" ", "").replace("\n", "")
            if cell_value in headers:
                cell_value = cell_value +  "_1"
            headers.append(cell_value)
            break
print(headers)
# 连接到 SQLite 数据库（如果数据库文件不存在，会自动创建）
conn = sqlite3.connect('my_database.db')  # 你可以替换为你想要的数据库文件名
cursor = conn.cursor()
columns_sql = ', '.join([f'"{col}" TEXT' for col in headers])
create_table_sql = f"""
CREATE TABLE IF NOT EXISTS 招聘信息 (
    {columns_sql}
);
"""
# 执行创建表格的 SQL 语句
cursor.execute(create_table_sql)

# 提交事务
conn.commit()


print("中文表格已成功创建！")
for i in range(7, row_count):
    rowList = []
    for j in range(1, column_count + 1):
        cell = chr(65+ j - 1) + str(i)
        cell_value = sheet[cell].value
        rowList.append(cell_value)
    insert_sql = '''
    INSERT INTO 招聘信息 (
        岗位代码, 用人单位序号, 用人单位名称, 岗位类别, 岗位名称, 从事工作, 招考数量, 
        入围比例, 来源类别, 学历, 学位, 所学专业, 考试专业科目, 高校毕业生, 社会人才, 
        高校毕业生_1, 社会人才_1, 其他条件, 工作地点, 咨询电话
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    '''
    cursor.execute(insert_sql, rowList)
    conn.commit()
    print("数据库已成功插入")
    print(rowList)
